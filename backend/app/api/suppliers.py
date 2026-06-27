"""Supplier management API routes."""

from typing import Optional, List
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from app.utils.database import get_db
from app.models import Supplier
import openpyxl
from openpyxl import Workbook
import io
import os
import xlrd
import tempfile
from urllib.parse import quote

router = APIRouter(prefix="/suppliers", tags=["Supplier Management"])


class SupplierCreate(BaseModel):
    code: str
    name: str
    contact_name: Optional[str] = None
    contact_phone: Optional[str] = None
    address: Optional[str] = None


class SupplierResponse(BaseModel):
    id: int
    code: str
    name: str
    contact_name: Optional[str] = None
    contact_phone: Optional[str] = None
    address: Optional[str] = None
    active: int = 1


@router.get("", response_model=List[SupplierResponse])
async def list_suppliers(db: AsyncSession = Depends(get_db)):
    """List all active suppliers."""
    result = await db.execute(select(Supplier).where(Supplier.active == 1).order_by(Supplier.id))
    suppliers = result.scalars().all()
    return [SupplierResponse(
        id=s.id, code=s.code, name=s.name,
        contact_name=s.contact_name, contact_phone=s.contact_phone,
        address=s.address, active=s.active,
    ) for s in suppliers]


@router.get("/all", response_model=List[SupplierResponse])
async def list_all_suppliers(db: AsyncSession = Depends(get_db)):
    """List all suppliers (including disabled)."""
    result = await db.execute(select(Supplier).order_by(Supplier.id))
    suppliers = result.scalars().all()
    return [SupplierResponse(
        id=s.id, code=s.code, name=s.name,
        contact_name=s.contact_name, contact_phone=s.contact_phone,
        address=s.address, active=s.active,
    ) for s in suppliers]


@router.post("", response_model=SupplierResponse)
async def create_supplier(data: SupplierCreate, db: AsyncSession = Depends(get_db)):
    """Create a new supplier."""
    existing = await db.execute(select(Supplier).where(Supplier.code == data.code))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="供应商编码已存在")

    supplier = Supplier(
        code=data.code,
        name=data.name,
        contact_name=data.contact_name,
        contact_phone=data.contact_phone,
        address=data.address,
        active=1,
    )
    db.add(supplier)
    await db.commit()
    await db.refresh(supplier)
    return SupplierResponse(
        id=supplier.id, code=supplier.code, name=supplier.name,
        contact_name=supplier.contact_name, contact_phone=supplier.contact_phone,
        address=supplier.address, active=supplier.active,
    )


@router.put("/{supplier_id}", response_model=SupplierResponse)
async def update_supplier(
    supplier_id: int,
    data: SupplierCreate,
    db: AsyncSession = Depends(get_db),
):
    """Update a supplier."""
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="供应商不存在")

    if data.code != supplier.code:
        existing = await db.execute(select(Supplier).where(Supplier.code == data.code))
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="供应商编码已存在")

    supplier.code = data.code
    supplier.name = data.name
    supplier.contact_name = data.contact_name
    supplier.contact_phone = data.contact_phone
    supplier.address = data.address
    await db.commit()
    await db.refresh(supplier)
    return SupplierResponse(
        id=supplier.id, code=supplier.code, name=supplier.name,
        contact_name=supplier.contact_name, contact_phone=supplier.contact_phone,
        address=supplier.address, active=supplier.active,
    )


@router.delete("/{supplier_id}")
async def delete_supplier(supplier_id: int, db: AsyncSession = Depends(get_db)):
    """Soft delete a supplier."""
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="供应商不存在")

    supplier.active = 0
    await db.commit()
    return {"status": "ok", "message": "供应商已删除", "supplier_id": supplier_id}


@router.get("/template")
async def download_supplier_template():
    """Download supplier import template (.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "供应商模板"
    headers = ["供应商编码", "供应商名称", "联系人", "联系电话", "地址"]
    ws.append(headers)
    ws.append(["SUP-001", "深圳电子元器件有限公司", "张三", "13800138000", "深圳市南山区科技园"])
    ws.append(["SUP-002", "上海芯片科技公司", "李四", "13900139000", "上海市浦东新区张江高科"])
    widths = [18, 28, 12, 16, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('供应商导入模板.xlsx')}"},
    )


@router.post("/upload")
async def upload_suppliers(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload suppliers from Excel template.

    Format: 供应商编码, 供应商名称, 联系人, 联系电话, 地址
    Duplicate supplier codes will be skipped.
    """
    content = await file.read()
    filename = file.filename or ""

    # Read rows from .xls or .xlsx
    rows = []
    if filename.lower().endswith(".xls"):
        with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            wb = xlrd.open_workbook(tmp_path)
            ws = wb.sheet_by_index(0)
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    row.append(ws.cell_value(r, c))
                rows.append(row)
        finally:
            os.unlink(tmp_path)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row) if row else [])

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="模板至少需要标题行和一行数据")

    total = 0
    imported = 0
    skipped = 0

    for row in rows[1:]:
        if len(row) < 2:
            continue
        code = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        contact_name = str(row[2] or "").strip() if len(row) > 2 else ""
        contact_phone = str(row[3] or "").strip() if len(row) > 3 else ""
        address = str(row[4] or "").strip() if len(row) > 4 else ""

        if not code or not name:
            continue

        total += 1

        # Check if supplier code already exists
        existing = await db.execute(
            select(Supplier).where(Supplier.code == code)
        )
        existing_sup = existing.scalar_one_or_none()
        if existing_sup:
            if existing_sup.active == 1:
                skipped += 1
                continue
            else:
                # Reactivate and update
                existing_sup.active = 1
                existing_sup.name = name
                existing_sup.contact_name = contact_name if contact_name else existing_sup.contact_name
                existing_sup.contact_phone = contact_phone if contact_phone else existing_sup.contact_phone
                existing_sup.address = address if address else existing_sup.address
                imported += 1
                continue

        supplier = Supplier(
            code=code,
            name=name,
            contact_name=contact_name or None,
            contact_phone=contact_phone or None,
            address=address or None,
            active=1,
        )
        db.add(supplier)
        imported += 1

    await db.commit()

    return {
        "total": total,
        "imported": imported,
        "skipped": skipped,
    }
