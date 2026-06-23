"""Material management API routes."""

from typing import Optional, List
from sqlalchemy import select, or_, delete, text, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from app.schemas import (
    MaterialCreate,
    MaterialUpdate,
    MaterialResponse,
    MaterialUploadResponse,
    CustomerMaterialMappingCreate,
    CustomerMaterialMappingUpdate,
    CustomerMaterialMappingResponse,
)
from app.utils.database import get_db
from app.models import MaterialMaster, MaterialCategory, CustomerMaterialMapping, Customer
import openpyxl
from openpyxl import Workbook
import io
import os
import re
import xlrd
import tempfile
from urllib.parse import quote

router = APIRouter(prefix="/materials", tags=["Material Management"])


@router.get("/template")
async def download_material_template():
    """Download material master import template (.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "物料主数据模板"
    headers = ["料号", "品名", "规格", "单位", "类别"]
    ws.append(headers)
    ws.append(["MAT-001", "贴片电阻", "100Ω,1/16W,J,0402,卷带", "PCS", "贴片电阻"])
    ws.append(["MAT-002", "贴片电容", "100nF,16V,K,X7R,0402,卷带", "PCS", "贴片电容"])
    widths = [22, 18, 40, 10, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('物料主数据导入模板.xlsx')}"},
    )


@router.post("/upload", response_model=MaterialUploadResponse)
async def upload_materials(
    file: UploadFile = File(...),
    customer_code: str = None,
    db: AsyncSession = Depends(get_db),
):
    """Upload material master data from Excel template.
    
    Format (Qixin template): 料号, 品名, 规格, 单位, 类别
    Duplicate material codes will be skipped (not overwritten).
    """
    if not customer_code:
        raise HTTPException(status_code=400, detail="客户编码不能为空")

    customer_result = await db.execute(select(Customer).where(Customer.code == customer_code))
    customer = customer_result.scalar_one_or_none()
    if customer is None:
        raise HTTPException(status_code=400, detail="客户不存在")

    content = await file.read()
    filename = file.filename or ""

    # Read rows from .xls or .xlsx
    rows = []
    if filename.lower().endswith(".xls"):
        with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            wb = xlrd.open_workbook(tmp_path)
            ws = wb.sheet_by_index(0)
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    row.append(ws.cell_value(r, c))
                rows.append(row)
        finally:
            os.unlink(tmp_path)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row) if row else [])

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="模板至少需要标题行和一行数据")

    # Parse: skip header row (index 0)
    total = 0
    imported = 0
    skipped = 0
    category_cache = {}

    # Customer prefix for material codes
    customer_prefix = customer.code + "-"

    for row in rows[1:]:
        if len(row) < 3:
            continue
        raw_code = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        spec = str(row[2] or "").strip()
        unit = str(row[3] or "PCS").strip() if len(row) > 3 else "PCS"
        category_name = str(row[4] or "").strip() if len(row) > 4 else ""

        if not raw_code or not name:
            continue

        # Auto-prepend customer prefix (avoid double prefix)
        if not raw_code.startswith(customer_prefix):
            code = customer_prefix + raw_code
        else:
            code = raw_code

        total += 1

        # Check if material code already exists
        existing = await db.execute(
            select(MaterialMaster).where(
                MaterialMaster.customer_id == customer.id,
                MaterialMaster.code == code,
            )
        )
        existing_mat = existing.scalar_one_or_none()
        if existing_mat:
            if existing_mat.active == 1:
                # Active duplicate → skip
                skipped += 1
                continue
            else:
                # Soft-deleted → reactivate and update
                existing_mat.active = 1
                existing_mat.name = name
                existing_mat.spec = spec
                existing_mat.unit = unit
                if category_id is not None:
                    existing_mat.category_id = category_id
                imported += 1
                continue

        # Resolve category
        category_id = None
        if category_name:
            if category_name in category_cache:
                category_id = category_cache[category_name]
            else:
                cat_result = await db.execute(
                    select(MaterialCategory).where(
                        MaterialCategory.customer_id == customer.id,
                        MaterialCategory.name == category_name,
                    )
                )
                cat = cat_result.scalar_one_or_none()
                if not cat:
                    # Auto-create category
                    cat_code = re.sub(r'[^A-Za-z0-9\u4e00-\u9fff]', '', category_name)[:20]
                    cat = MaterialCategory(
                        customer_id=customer.id,
                        name=category_name,
                        code=cat_code or "general",
                    )
                    db.add(cat)
                    await db.flush()
                category_id = cat.id
                category_cache[category_name] = category_id

        mat = MaterialMaster(
            customer_id=customer.id,
            code=code,
            name=name,
            spec=spec,
            unit=unit,
            category_id=category_id,
            active=1,
        )
        db.add(mat)
        imported += 1

    await db.commit()

    categories_created = sum(
        1 for name, cid in category_cache.items()
        if cid is not None
    )

    return MaterialUploadResponse(
        total=total,
        imported=imported,
        skipped=skipped,
        categories_created=categories_created,
    )


@router.get("")
async def list_materials(
    customer_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    show_disabled: Optional[bool] = Query(False),
    page: Optional[int] = Query(None, ge=1, description="Page number (omit to disable pagination)"),
    page_size: int = Query(20, ge=1, le=500, description="Items per page"),
    db: AsyncSession = Depends(get_db),
):
    # Build base query
    base_query = select(MaterialMaster)
    if customer_id:
        base_query = base_query.where(MaterialMaster.customer_id == customer_id)
    if category_id:
        base_query = base_query.where(MaterialMaster.category_id == category_id)
    if keyword:
        kw = f"%{keyword}%"
        base_query = base_query.where(
            or_(
                MaterialMaster.code.ilike(kw),
                MaterialMaster.name.ilike(kw),
                MaterialMaster.spec.ilike(kw)
            )
        )
    if not show_disabled:
        base_query = base_query.where(MaterialMaster.active == 1)
    base_query = base_query.order_by(MaterialMaster.code)

    # Get total count
    count_query = select(func.count()).select_from(base_query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar()

    # Apply pagination (only if page is specified; otherwise return all)
    if page is not None:
        query = base_query.offset((page - 1) * page_size).limit(page_size)
    else:
        query = base_query

    result = await db.execute(query)
    materials = result.scalars().all()

    return {
        "data": [
            MaterialResponse(
                id=m.id, code=m.code, name=m.name,
                spec=m.spec, unit=m.unit, qty_per_pallet=m.qty_per_pallet,
                active=m.active,
            ) for m in materials
        ],
        "total": total,
        "page": page if page is not None else 1,
        "page_size": page_size,
    }


@router.post("")
async def create_material(
    data: MaterialCreate,
    db: AsyncSession = Depends(get_db),
):
    existing = await db.execute(
        select(MaterialMaster).where(MaterialMaster.code == data.code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="物料编码已存在")
    material = MaterialMaster(
        code=data.code, name=data.name, spec=data.spec,
        unit=data.unit or "个",
        category_id=data.category_id, qty_per_pallet=data.qty_per_pallet,
        barcode_pattern=data.barcode_pattern, active=1,
        customer_id=1,
    )
    db.add(material)
    await db.commit()
    await db.refresh(material)
    return MaterialResponse(
        id=material.id, code=material.code, name=material.name,
        spec=material.spec, unit=material.unit,
        qty_per_pallet=material.qty_per_pallet,
        active=material.active,
    )


# ═══════════════════════════════════════════════
# Customer Material Mapping
# ═══════════════════════════════════════════════
# NOTE: These MUST be defined before /{material_id} routes to avoid
# FastAPI path conflict — "/mappings" would otherwise be captured
# as {material_id}=="mappings" and fail validation.

@router.get("/mappings", response_model=List[CustomerMaterialMappingResponse])
async def list_mappings(
    customer_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """List customer-material mappings, optionally filtered by customer."""
    query = select(
        CustomerMaterialMapping,
        MaterialMaster.code,
        MaterialMaster.name,
        Customer.name,
    ).join(
        MaterialMaster,
        CustomerMaterialMapping.internal_material_id == MaterialMaster.id,
    ).join(
        Customer,
        CustomerMaterialMapping.customer_id == Customer.id,
    )
    if customer_id:
        query = query.where(CustomerMaterialMapping.customer_id == customer_id)
    query = query.order_by(CustomerMaterialMapping.customer_id, CustomerMaterialMapping.customer_material_code)
    result = await db.execute(query)
    rows = result.all()
    return [
        CustomerMaterialMappingResponse(
            id=row[0].id,
            customer_id=row[0].customer_id,
            customer_material_code=row[0].customer_material_code,
            internal_material_id=row[0].internal_material_id,
            internal_material_code=row[1],
            internal_material_name=row[2],
            customer_name=row[3],
            active=row[0].active,
            created_at=row[0].created_at,
        )
        for row in rows
    ]


@router.post("/mappings", response_model=CustomerMaterialMappingResponse)
async def create_mapping(
    data: CustomerMaterialMappingCreate,
    db: AsyncSession = Depends(get_db),
):
    """Create a new customer-material mapping."""
    cust_result = await db.execute(
        select(Customer).where(Customer.id == data.customer_id)
    )
    customer = cust_result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    mat_result = await db.execute(
        select(MaterialMaster).where(MaterialMaster.id == data.internal_material_id)
    )
    material = mat_result.scalar_one_or_none()
    if not material:
        raise HTTPException(status_code=404, detail="物料不存在")

    dup = await db.execute(
        select(CustomerMaterialMapping).where(
            CustomerMaterialMapping.customer_id == data.customer_id,
            CustomerMaterialMapping.customer_material_code == data.customer_material_code,
        )
    )
    if dup.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail=f"该客户的物料编码 '{data.customer_material_code}' 已存在映射",
        )

    mapping = CustomerMaterialMapping(
        customer_id=data.customer_id,
        customer_material_code=data.customer_material_code,
        internal_material_id=data.internal_material_id,
        active=1,
    )
    db.add(mapping)
    await db.commit()
    await db.refresh(mapping)

    return CustomerMaterialMappingResponse(
        id=mapping.id,
        customer_id=mapping.customer_id,
        customer_material_code=mapping.customer_material_code,
        internal_material_id=mapping.internal_material_id,
        internal_material_code=material.code,
        internal_material_name=material.name,
        customer_name=customer.name,
        active=mapping.active,
        created_at=mapping.created_at,
    )


@router.put("/mappings/{mapping_id}", response_model=CustomerMaterialMappingResponse)
async def update_mapping(
    mapping_id: int,
    data: CustomerMaterialMappingUpdate,
    db: AsyncSession = Depends(get_db),
):
    """Update a customer-material mapping."""
    result = await db.execute(
        select(CustomerMaterialMapping).where(CustomerMaterialMapping.id == mapping_id)
    )
    mapping = result.scalar_one_or_none()
    if not mapping:
        raise HTTPException(status_code=404, detail="映射不存在")

    if data.customer_material_code is not None:
        mapping.customer_material_code = data.customer_material_code
    if data.internal_material_id is not None:
        mapping.internal_material_id = data.internal_material_id
    if data.active is not None:
        mapping.active = data.active

    await db.commit()
    await db.refresh(mapping)

    mat_result = await db.execute(
        select(MaterialMaster).where(MaterialMaster.id == mapping.internal_material_id)
    )
    material = mat_result.scalar_one_or_none()
    cust_result = await db.execute(
        select(Customer).where(Customer.id == mapping.customer_id)
    )
    customer = cust_result.scalar_one_or_none()

    return CustomerMaterialMappingResponse(
        id=mapping.id,
        customer_id=mapping.customer_id,
        customer_material_code=mapping.customer_material_code,
        internal_material_id=mapping.internal_material_id,
        internal_material_code=material.code if material else "",
        internal_material_name=material.name if material else "",
        customer_name=customer.name if customer else "",
        active=mapping.active,
        created_at=mapping.created_at,
    )


@router.delete("/mappings/{mapping_id}")
async def delete_mapping(
    mapping_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Soft-delete a customer-material mapping."""
    result = await db.execute(
        select(CustomerMaterialMapping).where(CustomerMaterialMapping.id == mapping_id)
    )
    mapping = result.scalar_one_or_none()
    if not mapping:
        raise HTTPException(status_code=404, detail="映射不存在")
    mapping.active = 0
    await db.commit()
    return {"status": "ok", "message": "映射已禁用"}


@router.get("/{material_id}")
async def get_material(
    material_id: int,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(MaterialMaster).where(MaterialMaster.id == material_id)
    )
    m = result.scalar_one_or_none()
    if not m:
        raise HTTPException(status_code=404, detail="物料不存在")
    return MaterialResponse(
        id=m.id, code=m.code, name=m.name,
        spec=m.spec, unit=m.unit, qty_per_pallet=m.qty_per_pallet,
        active=m.active,
    )


@router.put("/{material_id}")
async def update_material(
    material_id: int,
    data: MaterialUpdate,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(MaterialMaster).where(MaterialMaster.id == material_id)
    )
    m = result.scalar_one_or_none()
    if not m:
        raise HTTPException(status_code=404, detail="物料不存在")
    if data.code is not None:
        m.code = data.code
    if data.name is not None:
        m.name = data.name
    if data.spec is not None:
        m.spec = data.spec
    if data.unit is not None:
        m.unit = data.unit
    if data.category_id is not None:
        m.category_id = data.category_id
    if data.qty_per_pallet is not None:
        m.qty_per_pallet = data.qty_per_pallet
    if data.barcode_pattern is not None:
        m.barcode_pattern = data.barcode_pattern
    if data.active is not None:
        m.active = data.active
    await db.commit()
    await db.refresh(m)
    return MaterialResponse(
        id=m.id, code=m.code, name=m.name,
        spec=m.spec, unit=m.unit, qty_per_pallet=m.qty_per_pallet,
        active=m.active,
    )


@router.delete("/{material_id}")
async def delete_material(
    material_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Permanently delete a material (hard delete).

    - If only customer_material_mappings references exist, auto-deletes them first,
      then deletes the material.
    - If other constraints exist (inventory, receipts, issues, BOMs, transactions),
      returns an error with details.
    """
    result = await db.execute(
        select(MaterialMaster).where(MaterialMaster.id == material_id)
    )
    m = result.scalar_one_or_none()
    if not m:
        raise HTTPException(status_code=404, detail="物料不存在")

    # Check FK references
    ref_tables = [
        ("inventory_reels", "material_id", "库存记录"),
        ("receipt_reels", "material_id", "收料记录"),
        ("issue_detail", "material_id", "发料明细"),
        ("transactions", "material_id", "交易记录"),
        ("boms", "product_material_id", "BOM（产品物料）"),
        ("bom_items", "material_id", "BOM物料项"),
        ("bom_alternatives", "alternative_material_id", "BOM替代料"),
        ("customer_material_mappings", "internal_material_id", "客户料号映射"),
    ]

    # Dynamically discover additional FK references from PostgreSQL catalog
    fk_discovery_sql = text("""
        SELECT
            pgc.conrelid::regclass::text AS ref_table,
            a.attname AS ref_column,
            pgc.conname AS constraint_name
        FROM pg_constraint pgc
        JOIN pg_attribute a
            ON a.attnum = ANY(pgc.conkey)
            AND a.attrelid = pgc.conrelid
        WHERE pgc.contype = 'f'
          AND pgc.conrelid::regclass::text != 'material_master'
          AND pgc.confrelid::regclass::text = 'material_master'
          AND pgc.conkey[1] = a.attnum
    """)
    try:
        fk_result = await db.execute(fk_discovery_sql)
        fk_rows = fk_result.fetchall()
        for row in fk_rows:
            ref_table = row[0]
            ref_column = row[1]
            constraint_name = row[2]
            if not any(r[0] == ref_table for r in ref_tables):
                ref_tables.append(
                    (ref_table, ref_column, f"外键约束:{constraint_name}")
                )
    except Exception:
        pass

    referenced_by = []
    has_mapping_ref = False
    has_other_ref = False

    for table, fk_column, label in ref_tables:
        try:
            count_result = await db.execute(
                text(f"SELECT COUNT(*) FROM {table} WHERE {fk_column} = :mid"),
                {"mid": material_id},
            )
            count = count_result.scalar()
            if count and count > 0:
                ref_info = f"{label}({count}条)"
                referenced_by.append(ref_info)
                if table == "customer_material_mappings":
                    has_mapping_ref = True
                else:
                    has_other_ref = True
        except Exception:
            pass

    if not referenced_by:
        # No references → safe to hard delete
        await db.execute(
            delete(MaterialMaster).where(MaterialMaster.id == material_id)
        )
        await db.commit()
        return {
            "status": "ok",
            "message": f"物料 '{m.code}' 已永久删除",
            "deleted": True,
        }

    if not has_other_ref and has_mapping_ref:
        # Only customer_material_mappings → auto-cascade: delete mappings first
        try:
            await db.execute(
                text(
                    "DELETE FROM customer_material_mappings "
                    "WHERE internal_material_id = :mid"
                ),
                {"mid": material_id},
            )
            await db.execute(
                delete(MaterialMaster).where(MaterialMaster.id == material_id)
            )
            await db.commit()
            return {
                "status": "ok",
                "message": (
                    f"物料 '{m.code}' 已永久删除"
                    f"（自动清理 {referenced_by[0]}）"
                ),
                "deleted": True,
                "cascaded_mappings": True,
            }
        except Exception as e:
            await db.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"删除失败（自动级联客户料号时异常）: {type(e).__name__}: {str(e)}",
            )

    # Has other references → cannot delete
    ref_detail = "、".join(referenced_by)
    raise HTTPException(
        status_code=409,
        detail=(
            f"物料 '{m.code}' 无法删除，已被以下模块引用：{ref_detail}。"
            f"请先清理相关引用记录后再试。"
        ),
    )


# ═══════════════════════════════════════════════
# Batch operations
# ═══════════════════════════════════════════════

@router.post("/batch-delete")
async def batch_delete_materials(
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
):
    """Batch soft-delete materials."""
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="物料ID列表不能为空")
    result = await db.execute(
        select(MaterialMaster).where(MaterialMaster.id.in_(ids))
    )
    materials = result.scalars().all()
    for m in materials:
        m.active = 0
    await db.commit()
    return {"status": "ok", "message": f"已批量禁用 {len(materials)} 个物料", "count": len(materials)}


@router.put("/batch-update")
async def batch_update_materials(
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
):
    """Batch update materials (e.g., change category for multiple materials)."""
    ids = data.get("ids", [])
    fields = data.get("fields", {})
    if not ids:
        raise HTTPException(status_code=400, detail="物料ID列表不能为空")
    result = await db.execute(
        select(MaterialMaster).where(MaterialMaster.id.in_(ids))
    )
    materials = result.scalars().all()
    for m in materials:
        if "name" in fields:
            m.name = fields["name"]
        if "spec" in fields:
            m.spec = fields["spec"]
        if "unit" in fields:
            m.unit = fields["unit"]
        if "category_id" in fields:
            m.category_id = fields["category_id"]
        if "qty_per_pallet" in fields:
            m.qty_per_pallet = fields["qty_per_pallet"]
    await db.commit()
    return {"status": "ok", "message": f"已批量更新 {len(materials)} 个物料", "count": len(materials)}


@router.post("/batch-delete-permanently")
async def batch_delete_materials_permanently(
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
):
    """Permanently delete materials (hard delete from database).

    Only deletes if the material is not referenced by any existing records.
    """
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="物料ID列表不能为空")

    try:
        return await _do_batch_delete_permanently(ids, db)
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"批量永久删除异常: {type(e).__name__}: {str(e)}",
        )


async def _do_batch_delete_permanently(
    ids: list[int],
    db: AsyncSession,
) -> dict:
    """Internal implementation of batch permanent delete."""
    ref_tables = [
        ("inventory_reels", "material_id", "库存记录"),
        ("receipt_reels", "material_id", "收料记录"),
        ("issue_detail", "material_id", "发料明细"),
        ("transactions", "material_id", "交易记录"),
        ("boms", "product_material_id", "BOM（产品物料）"),
        ("bom_items", "material_id", "BOM物料项"),
        ("bom_alternatives", "alternative_material_id", "BOM替代料"),
        ("customer_material_mappings", "internal_material_id", "客户物料映射"),
    ]

    # Discover additional FK references from PostgreSQL catalog
    fk_discovery_sql = text("""
        SELECT
            pgc.conrelid::regclass::text AS ref_table,
            a.attname AS ref_column,
            pgc.conname AS constraint_name
        FROM pg_constraint pgc
        JOIN pg_attribute a
            ON a.attnum = ANY(pgc.conkey)
            AND a.attrelid = pgc.conrelid
        WHERE pgc.contype = 'f'
          AND pgc.conrelid::regclass::text != 'material_master'
          AND pgc.confrelid::regclass::text = 'material_master'
          AND pgc.conkey[1] = a.attnum
    """)
    try:
        result = await db.execute(fk_discovery_sql)
        fk_rows = result.fetchall()
        for row in fk_rows:
            ref_table = row[0]
            ref_column = row[1]
            constraint_name = row[2]
            if not any(r[0] == ref_table for r in ref_tables):
                ref_tables.append(
                    (ref_table, ref_column, f"外键约束:{constraint_name}")
                )
    except Exception:
        pass

    deleted_ids = []
    skipped = []

    # Phase 1: Check references and collect safe-to-delete IDs
    safe_ids = []
    for mid in ids:
        referenced_by = []
        for table, fk_column, label in ref_tables:
            try:
                result = await db.execute(
                    text(
                        f"SELECT COUNT(*) FROM {table} WHERE {fk_column} = :mid"
                    ),
                    {"mid": mid},
                )
                count = result.scalar()
                if count and count > 0:
                    referenced_by.append(f"{label}({count}条)")
            except Exception:
                pass

        if referenced_by:
            mat_result = await db.execute(
                select(MaterialMaster.code).where(MaterialMaster.id == mid)
            )
            mat_code = mat_result.scalar_one_or_none() or f"ID={mid}"
            skipped.append(
                {"id": mid, "code": mat_code, "references": referenced_by}
            )
        else:
            safe_ids.append(mid)

    # Phase 2: Delete all safe IDs in one operation
    if safe_ids:
        try:
            await db.execute(
                delete(MaterialMaster).where(MaterialMaster.id.in_(safe_ids))
            )
            deleted_ids.extend(safe_ids)
        except IntegrityError:
            await db.rollback()
            # Fallback: try deleting one by one
            for mid in safe_ids:
                try:
                    await db.execute(
                        delete(MaterialMaster).where(MaterialMaster.id == mid)
                    )
                    deleted_ids.append(mid)
                except IntegrityError:
                    await db.rollback()
                    mat_result = await db.execute(
                        select(MaterialMaster.code).where(
                            MaterialMaster.id == mid
                        )
                    )
                    mat_code = (
                        mat_result.scalar_one_or_none() or f"ID={mid}"
                    )
                    skipped.append({
                        "id": mid,
                        "code": mat_code,
                        "references": ["外键约束拦截"],
                    })
                except Exception:
                    await db.rollback()

    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(
            status_code=500, detail="提交删除失败，请检查数据库连接"
        )

    msg_parts = []
    if deleted_ids:
        msg_parts.append(f"已永久删除 {len(deleted_ids)} 个物料")
    if skipped:
        msg_parts.append(
            f"{len(skipped)} 个物料因存在关联记录被跳过"
        )

    return {
        "status": "ok",
        "message": "；".join(msg_parts)
        if msg_parts
        else "未执行任何删除操作",
        "deleted": deleted_ids,
        "deleted_count": len(deleted_ids),
        "skipped": skipped,
    }
