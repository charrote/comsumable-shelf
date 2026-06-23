"""BOM API routes."""

from datetime import datetime
from typing import Optional, List
from sqlalchemy import select, delete, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from app.schemas import (
    BomCreateRequest, BomUpdateRequest, BomListItem, BomDetailResponse,
    BomItemSchema, BomAlternativeSchema, BomUploadResponse, BomGenerateIssueRequest,
)
from app.utils.database import get_db
from app.models import Bom, BomItem, BomAlternative, MaterialMaster, Customer
from app.config import settings
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse
import os
import io
import re
import xlrd
from urllib.parse import quote

router = APIRouter(prefix="/bom", tags=["BOM"])


@router.get("/template")
async def download_bom_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM模板"
    headers = ["产品编码", "物料编码", "数量", "单位", "父级物料编码", "替代物料编码", "替代优先级", "替代百分比"]
    ws.append(headers)
    ws.append(["PROD-001", "MAT-001", 10, "盘", "", "", "", ""])
    ws.append(["PROD-001", "MAT-002", 5, "个", "MAT-001", "", "", ""])
    ws.append(["PROD-001", "MAT-003", 2, "套", "MAT-001", "MAT-004", 1, 100])
    widths = [15, 15, 10, 10, 18, 18, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('BOM导入模板.xlsx')}"},
    )


@router.get("/template-qixin")
async def download_bom_template_qixin():
    """Download Qixin-format BOM import template (.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM模板"
    # Row 1: product description
    ws.merge_cells("A1:F1")
    ws["A1"] = "PROD-001 (产品名称)         生产数量: 1000"
    # Row 2: headers
    headers = ["替代关系", "料号", "品名", "物料规格", "组成量", "位置号"]
    ws.append(headers)
    # Row 3+: sample data
    ws.append(["BOM子件", "MAT-001", "电阻", "100Ω,1/16W,J,0402,卷带", 10, "R1 R2 R3"])
    ws.append(["替代料", "MAT-002", "电阻(替代)", "100Ω,1/16W,F,0402,卷带", 10, ""])
    ws.append(["BOM子件", "MAT-003", "电容", "100nF,16V,K,X7R,0402,卷带", 5, "C1 C2"])
    ws.append(["BOM子件", "MAT-004", "IC主控", "MT9256, BGA479, 盘装", 1, "U1"])
    ws.append(["替代料", "MAT-005", "IC主控(替代)", "MT9255, BGA479, 盘装", 1, ""])
    # Column widths
    widths = [12, 22, 18, 40, 10, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('七鑫BOM导入模板.xlsx')}"},
    )


@router.get("", response_model=List[BomListItem])
async def list_boms(
    customer_id: Optional[int] = None,
    product_code: Optional[str] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    query = select(Bom).options(
        selectinload(Bom.product_material),
        selectinload(Bom.customer),
    )
    if customer_id is not None:
        query = query.where(Bom.customer_id == customer_id)
    if product_code:
        query = query.join(MaterialMaster, Bom.product_material_id == MaterialMaster.id).where(
            MaterialMaster.code.ilike(f"%{product_code}%")
        )
    if status:
        query = query.where(Bom.status == status)
    query = query.order_by(Bom.created_at.desc())
    result = await db.execute(query)
    boms = result.scalars().all()

    items = []
    for bom in boms:
        item_count_result = await db.execute(
            select(func.count(BomItem.id)).where(BomItem.bom_id == bom.id)
        )
        item_count = item_count_result.scalar() or 0
        items.append(BomListItem(
            id=bom.id,
            customer_id=bom.customer_id,
            customer_name=bom.customer.name if bom.customer else None,
            product_material_id=bom.product_material_id,
            product_code=bom.product_material.code if bom.product_material else None,
            product_name=bom.product_material.name if bom.product_material else None,
            version=bom.version,
            status=bom.status,
            description=bom.description,
            item_count=item_count,
            created_at=bom.created_at,
            updated_at=bom.updated_at,
        ))
    return items


@router.post("", response_model=BomDetailResponse)
async def create_bom(
    data: BomCreateRequest,
    db: AsyncSession = Depends(get_db),
):
    product = await db.get(MaterialMaster, data.product_material_id)
    if not product:
        raise HTTPException(status_code=400, detail="产品物料不存在")
    existing = await db.execute(
        select(Bom).where(
            Bom.customer_id == data.customer_id,
            Bom.product_material_id == data.product_material_id,
            Bom.version == data.version,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"版本 {data.version} 已存在")
    bom = Bom(
        customer_id=data.customer_id,
        product_material_id=data.product_material_id,
        version=data.version,
        status="draft",
        description=data.description,
    )
    db.add(bom)
    await db.commit()
    await db.refresh(bom)
    return await _get_bom_detail(db, bom.id)


@router.get("/{bom_id}", response_model=BomDetailResponse)
async def get_bom(bom_id: int, db: AsyncSession = Depends(get_db)):
    return await _get_bom_detail(db, bom_id)


async def _get_bom_detail(db: AsyncSession, bom_id: int) -> BomDetailResponse:
    result = await db.execute(
        select(Bom).where(Bom.id == bom_id).options(
            selectinload(Bom.product_material),
            selectinload(Bom.customer),
        )
    )
    bom = result.scalar_one_or_none()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")

    items_result = await db.execute(
        select(BomItem).where(BomItem.bom_id == bom_id).options(
            selectinload(BomItem.material),
            selectinload(BomItem.alternatives).selectinload(BomAlternative.alternative_material),
        )
    )
    all_items = items_result.scalars().all()

    def build_tree(parent_id: Optional[int]) -> List[BomItemSchema]:
        children = []
        for item in all_items:
            if item.parent_id == parent_id:
                alternatives = [
                    BomAlternativeSchema(
                        id=alt.id,
                        alternative_material_id=alt.alternative_material_id,
                        alternative_material_code=alt.alternative_material.code if alt.alternative_material else None,
                        alternative_material_name=alt.alternative_material.name if alt.alternative_material else None,
                        priority=alt.priority,
                        percentage=alt.percentage,
                    )
                    for alt in sorted(item.alternatives, key=lambda a: a.priority)
                ]
                child = BomItemSchema(
                    id=item.id,
                    parent_id=item.parent_id,
                    material_id=item.material_id,
                    material_code=item.material.code if item.material else None,
                    material_name=item.material.name if item.material else None,
                    material_unit=item.material.unit if item.material else None,
                    quantity=item.quantity,
                    position=item.position,
                    remark=item.remark,
                    alternatives=alternatives,
                    children=build_tree(item.id),
                )
                children.append(child)
        return sorted(children, key=lambda x: x.position)

    return BomDetailResponse(
        id=bom.id,
        customer_id=bom.customer_id,
        customer_name=bom.customer.name if bom.customer else None,
        product_material_id=bom.product_material_id,
        product_code=bom.product_material.code if bom.product_material else None,
        product_name=bom.product_material.name if bom.product_material else None,
        version=bom.version,
        status=bom.status,
        description=bom.description,
        items=build_tree(None),
        created_at=bom.created_at,
        updated_at=bom.updated_at,
    )


@router.put("/{bom_id}", response_model=BomDetailResponse)
async def update_bom(
    bom_id: int,
    data: BomUpdateRequest,
    db: AsyncSession = Depends(get_db),
):
    bom = await db.get(Bom, bom_id)
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")
    if data.version is not None:
        bom.version = data.version
    if data.status is not None:
        bom.status = data.status
    if data.description is not None:
        bom.description = data.description
    await db.commit()
    await db.refresh(bom)
    return await _get_bom_detail(db, bom_id)


@router.delete("/{bom_id}")
async def delete_bom(bom_id: int, db: AsyncSession = Depends(get_db)):
    bom = await db.get(Bom, bom_id)
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")

    # Check if any IssueOrder references this BOM
    from app.models import IssueOrder
    issue_check = await db.execute(
        select(IssueOrder).where(IssueOrder.bom_id == bom_id).limit(1)
    )
    if issue_check.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail=f"此BOM已被发料单引用，无法删除。请先删除相关的发料单（IssueOrder）后再试。"
        )

    await db.delete(bom)
    await db.commit()
    return {"message": "BOM已删除"}


@router.post("/{bom_id}/items", response_model=BomItemSchema)
async def add_bom_item(
    bom_id: int,
    data: BomItemSchema,
    db: AsyncSession = Depends(get_db),
):
    bom = await db.get(Bom, bom_id)
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")
    if data.parent_id:
        parent = await db.get(BomItem, data.parent_id)
        if not parent or parent.bom_id != bom_id:
            raise HTTPException(status_code=400, detail="父级物料不存在或不属于此BOM")

    item = BomItem(
        bom_id=bom_id,
        parent_id=data.parent_id,
        material_id=data.material_id,
        quantity=data.quantity,
        position=data.position,
        remark=data.remark,
    )
    db.add(item)
    await db.flush()

    for alt in data.alternatives:
        bom_alt = BomAlternative(
            bom_item_id=item.id,
            alternative_material_id=alt.alternative_material_id,
            priority=alt.priority,
            percentage=alt.percentage,
        )
        db.add(bom_alt)
    await db.commit()
    await db.refresh(item)

    return await _get_item_schema(db, item.id)


async def _get_item_schema(db: AsyncSession, item_id: int) -> BomItemSchema:
    result = await db.execute(
        select(BomItem).where(BomItem.id == item_id).options(
            selectinload(BomItem.material),
            selectinload(BomItem.alternatives).selectinload(BomAlternative.alternative_material),
        )
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="BOM明细不存在")

    alternatives = [
        BomAlternativeSchema(
            id=alt.id,
            alternative_material_id=alt.alternative_material_id,
            alternative_material_code=alt.alternative_material.code if alt.alternative_material else None,
            alternative_material_name=alt.alternative_material.name if alt.alternative_material else None,
            priority=alt.priority,
            percentage=alt.percentage,
        )
        for alt in sorted(item.alternatives, key=lambda a: a.priority)
    ]
    return BomItemSchema(
        id=item.id,
        parent_id=item.parent_id,
        material_id=item.material_id,
        material_code=item.material.code if item.material else None,
        material_name=item.material.name if item.material else None,
        material_unit=item.material.unit if item.material else None,
        quantity=item.quantity,
        position=item.position,
        remark=item.remark,
        alternatives=alternatives,
        children=[],
    )


@router.put("/{bom_id}/items/{item_id}", response_model=BomItemSchema)
async def update_bom_item(
    bom_id: int,
    item_id: int,
    data: BomItemSchema,
    db: AsyncSession = Depends(get_db),
):
    item = await db.get(BomItem, item_id)
    if not item or item.bom_id != bom_id:
        raise HTTPException(status_code=404, detail="BOM明细不存在")

    item.material_id = data.material_id
    item.quantity = data.quantity
    item.position = data.position
    item.remark = data.remark
    if data.parent_id:
        parent = await db.get(BomItem, data.parent_id)
        if not parent or parent.bom_id != bom_id:
            raise HTTPException(status_code=400, detail="父级物料不存在")
    item.parent_id = data.parent_id

    await db.execute(delete(BomAlternative).where(BomAlternative.bom_item_id == item_id))
    for alt in data.alternatives:
        bom_alt = BomAlternative(
            bom_item_id=item_id,
            alternative_material_id=alt.alternative_material_id,
            priority=alt.priority,
            percentage=alt.percentage,
        )
        db.add(bom_alt)
    await db.commit()
    return await _get_item_schema(db, item_id)


@router.delete("/{bom_id}/items/{item_id}")
async def delete_bom_item(
    bom_id: int,
    item_id: int,
    db: AsyncSession = Depends(get_db),
):
    item = await db.get(BomItem, item_id)
    if not item or item.bom_id != bom_id:
        raise HTTPException(status_code=404, detail="BOM明细不存在")
    await db.delete(item)
    await db.commit()
    return {"message": "BOM明细已删除"}


@router.post("/upload", response_model=BomUploadResponse)
async def upload_bom(
    file: UploadFile = File(...),
    customer_code: str = None,
    version: str = "1.0",
    db: AsyncSession = Depends(get_db),
):
    if not customer_code:
        raise HTTPException(status_code=400, detail="客户编码不能为空")

    customer_result = await db.execute(select(Customer).where(Customer.code == customer_code))
    customer = customer_result.scalar_one_or_none()
    if customer is None:
        raise HTTPException(status_code=400, detail="客户不存在")

    data_dir = os.path.join(os.path.dirname(__file__), "..", "data", "bom")
    os.makedirs(data_dir, exist_ok=True)
    file_path = os.path.join(data_dir, file.filename)
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    rows_data = []
    product_codes = set()
    material_codes = set()
    alternates_found = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 3:
            continue
        product_code = str(row[0] or "").strip()
        material_code = str(row[1] or "").strip()
        quantity = float(row[2] or 0)
        unit = str(row[3] or "盘").strip() if len(row) > 3 else "盘"
        parent_code = str(row[4] or "").strip() if len(row) > 4 else ""
        alt_code = str(row[5] or "").strip() if len(row) > 5 else ""
        alt_priority = int(row[6] or 1) if len(row) > 6 else 1
        alt_percentage = float(row[7] or 100) if len(row) > 7 else 100.0

        if not product_code or not material_code or quantity <= 0:
            continue

        product_codes.add(product_code)
        material_codes.add(material_code)
        if alt_code:
            alternates_found += 1
            material_codes.add(alt_code)

        rows_data.append({
            "product_code": product_code,
            "material_code": material_code,
            "quantity": quantity,
            "unit": unit,
            "parent_code": parent_code,
            "alt_code": alt_code,
            "alt_priority": alt_priority,
            "alt_percentage": alt_percentage,
        })

    auto_created = 0
    all_codes = product_codes | material_codes
    for code in all_codes:
        existing = await db.execute(
            select(MaterialMaster).where(
                MaterialMaster.customer_id == customer.id,
                MaterialMaster.code == code,
            )
        )
        if not existing.scalar_one_or_none():
            mat = MaterialMaster(
                customer_id=customer.id,
                code=code,
                name=code,
                unit="盘",
                active=1,
            )
            db.add(mat)
            auto_created += 1
    await db.flush()

    material_cache = {}
    for code in all_codes:
        result = await db.execute(
            select(MaterialMaster).where(
                MaterialMaster.customer_id == customer.id,
                MaterialMaster.code == code,
            )
        )
        mat = result.scalar_one_or_none()
        if mat:
            material_cache[code] = mat

    created_boms = {}
    item_code_to_id = {}

    for product_code in product_codes:
        product_mat = material_cache.get(product_code)
        if not product_mat:
            continue
        existing = await db.execute(
            select(Bom).where(
                Bom.customer_id == customer.id,
                Bom.product_material_id == product_mat.id,
                Bom.version == version,
            )
        )
        bom = existing.scalar_one_or_none()
        if not bom:
            bom = Bom(
                customer_id=customer.id,
                product_material_id=product_mat.id,
                version=version,
                status="draft",
            )
            db.add(bom)
            await db.flush()
        created_boms[product_code] = bom

    for row in rows_data:
        product_code = row["product_code"]
        bom = created_boms.get(product_code)
        if not bom:
            continue

        material = material_cache.get(row["material_code"])
        if not material:
            continue

        parent_id = None
        if row["parent_code"]:
            parent_id = item_code_to_id.get((product_code, row["parent_code"]))

        item = BomItem(
            bom_id=bom.id,
            parent_id=parent_id,
            material_id=material.id,
            quantity=row["quantity"],
            position=0,
        )
        db.add(item)
        await db.flush()
        item_code_to_id[(product_code, row["material_code"])] = item.id

        if row["alt_code"]:
            alt_material = material_cache.get(row["alt_code"])
            if alt_material:
                bom_alt = BomAlternative(
                    bom_item_id=item.id,
                    alternative_material_id=alt_material.id,
                    priority=row["alt_priority"],
                    percentage=row["alt_percentage"],
                )
                db.add(bom_alt)

    await db.commit()

    first_product = next(iter(product_codes), "")
    first_bom = created_boms.get(first_product)

    return BomUploadResponse(
        bom_id=first_bom.id if first_bom else 0,
        product_code=first_product,
        version=version,
        parsed=True,
        total_items=len(rows_data),
        unique_materials=len(material_codes),
        alternates_found=alternates_found,
        auto_created_count=auto_created,
    )


@router.post("/upload-qixin", response_model=BomUploadResponse)
async def upload_bom_qixin(
    file: UploadFile = File(...),
    customer_code: str = None,
    version: str = "1.0",
    db: AsyncSession = Depends(get_db),
):
    """Upload BOM using Qixin-format template.
    
    Format:
      Row 1: Product description (first word → product code)
      Row 2: Headers (替代关系, 料号, 品名, 物料规格, 组成量, 位置号)
      Row 3+: Data
        - 替代关系 = "BOM子件" (regular item) or "替代料" (alternative)
        - Alternatives are on separate rows directly below their parent
    """
    if not customer_code:
        raise HTTPException(status_code=400, detail="客户编码不能为空")

    customer_result = await db.execute(select(Customer).where(Customer.code == customer_code))
    customer = customer_result.scalar_one_or_none()
    if customer is None:
        raise HTTPException(status_code=400, detail="客户不存在")

    # Determine file type and read content
    content = await file.read()
    filename = file.filename or ""

    if filename.lower().endswith(".xls"):
        # Use xlrd for .xls
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            wb = xlrd.open_workbook(tmp_path)
            ws = wb.sheet_by_index(0)
            rows = []
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    row.append(ws.cell_value(r, c))
                rows.append(row)
        finally:
            os.unlink(tmp_path)
    else:
        # Use openpyxl for .xlsx
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row) if row else [])

    if len(rows) < 3:
        raise HTTPException(status_code=400, detail="模板数据不足，至少需要标题行和一行数据")

    # Row 1: product description → extract product code
    product_desc = str(rows[0][0] or "").strip()
    product_code_match = re.match(r"^(\S+)", product_desc)
    product_code = product_code_match.group(1) if product_code_match else product_desc
    if not product_code or product_code in ("替代关系",):
        # Fallback: use a generated code based on content
        product_code = f"PROD-{datetime.now().strftime('%Y%m%d%H%M%S')}"

    # Row 2: headers (skip)

    # Parse data rows (from row index 2 onwards)
    row_data = []
    material_codes = set()
    alternates_found = 0

    for row in rows[2:]:
        if not row or len(row) < 2:
            continue
        alt_rel = str(row[0] or "").strip()
        mat_code = str(row[1] or "").strip()
        mat_name = str(row[2] or "").strip() if len(row) > 2 else ""
        mat_spec = str(row[3] or "").strip() if len(row) > 3 else ""
        qty = float(row[4]) if len(row) > 4 and row[4] else 0
        position = str(row[5] or "").strip() if len(row) > 5 else ""

        if not mat_code or qty <= 0:
            continue

        is_alternative = "替代料" in alt_rel
        if is_alternative:
            alternates_found += 1

        material_codes.add(mat_code)

        row_data.append({
            "is_alternative": is_alternative,
            "material_code": mat_code,
            "material_name": mat_name,
            "material_spec": mat_spec,
            "quantity": qty,
            "position": position,
        })

    if not row_data:
        raise HTTPException(status_code=400, detail="未能解析到有效的BOM数据行")

    # Auto-create materials
    auto_created = 0
    for code in material_codes:
        existing = await db.execute(
            select(MaterialMaster).where(
                MaterialMaster.customer_id == customer.id,
                MaterialMaster.code == code,
            )
        )
        if not existing.scalar_one_or_none():
            # Find name and spec from first occurrence
            name = code
            spec = ""
            for rd in row_data:
                if rd["material_code"] == code:
                    name = rd["material_name"] or code
                    spec = rd["material_spec"] or ""
                    break
            mat = MaterialMaster(
                customer_id=customer.id,
                code=code,
                name=name,
                spec=spec,
                unit="PCS",
                active=1,
            )
            db.add(mat)
            auto_created += 1
    await db.flush()

    # Build material cache
    material_cache = {}
    for code in material_codes:
        result = await db.execute(
            select(MaterialMaster).where(
                MaterialMaster.customer_id == customer.id,
                MaterialMaster.code == code,
            )
        )
        mat = result.scalar_one_or_none()
        if mat:
            material_cache[code] = mat

    # Ensure product material exists
    product_mat = material_cache.get(product_code)
    if not product_mat:
        product_mat = MaterialMaster(
            customer_id=customer.id,
            code=product_code,
            name=product_desc or product_code,
            unit="PCS",
            active=1,
        )
        db.add(product_mat)
        await db.flush()
        material_cache[product_code] = product_mat
        auto_created += 1

    # Create BOM
    existing_bom = await db.execute(
        select(Bom).where(
            Bom.customer_id == customer.id,
            Bom.product_material_id == product_mat.id,
            Bom.version == version,
        )
    )
    bom = existing_bom.scalar_one_or_none()
    if not bom:
        bom = Bom(
            customer_id=customer.id,
            product_material_id=product_mat.id,
            version=version,
            status="draft",
            description=product_desc,
        )
        db.add(bom)
        await db.flush()

    # Process data rows — create BOM items and alternatives
    last_bom_item_id = None
    item_count = 0

    for rd in row_data:
        material = material_cache.get(rd["material_code"])
        if not material:
            continue

        if not rd["is_alternative"]:
            # Create new BOM item
            item = BomItem(
                bom_id=bom.id,
                material_id=material.id,
                quantity=rd["quantity"],
                position=item_count,
                remark=rd["position"] if rd["position"] else None,
            )
            db.add(item)
            await db.flush()
            last_bom_item_id = item.id
            item_count += 1
        else:
            # Add alternative to last BOM item
            if last_bom_item_id is not None:
                bom_alt = BomAlternative(
                    bom_item_id=last_bom_item_id,
                    alternative_material_id=material.id,
                    priority=1,
                    percentage=100.0,
                )
                db.add(bom_alt)

    await db.commit()

    return BomUploadResponse(
        bom_id=bom.id,
        product_code=product_code,
        version=version,
        parsed=True,
        total_items=item_count,
        unique_materials=len(material_codes),
        alternates_found=alternates_found,
        auto_created_count=auto_created,
    )


@router.post("/{bom_id}/generate-issue")
async def generate_issue_from_bom(
    bom_id: int,
    data: BomGenerateIssueRequest,
    db: AsyncSession = Depends(get_db),
):
    bom = await db.get(Bom, bom_id)
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")
    return {"message": "功能开发中", "bom_id": bom_id}


@router.get("/{bom_id}/export")
async def export_bom(
    bom_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Export BOM structure to Excel (.xlsx)."""
    detail = await _get_bom_detail(db, bom_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"BOM-{detail.product_code}"

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{detail.product_code} ({detail.product_name or ''})    版本: {detail.version}    状态: {detail.status}"
    ws["A1"].font = Font(bold=True, size=12)

    # Header row
    headers = ["层级", "物料编码", "物料名称", "数量", "单位", "位置/备注", "替代料"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Data rows - flatten tree
    def flatten_items(items: list, level: int = 0):
        rows = []
        for item in items:
            alt_str = "; ".join([
                f"{a.alternative_material_code}(优先级:{a.priority},{a.percentage}%)"
                for a in item.alternatives
            ]) if item.alternatives else ""
            rows.append([
                level,  # 层级
                item.material_code or "",
                item.material_name or "",
                item.quantity,
                item.material_unit or "",
                item.remark or "",
                alt_str,
            ])
            if item.children:
                rows.extend(flatten_items(item.children, level + 1))
        return rows

    data_rows = flatten_items(detail.items)
    for row in data_rows:
        ws.append(row)

    # Column widths
    widths = [8, 22, 30, 10, 8, 30, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"BOM-{detail.product_code}-{detail.version}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )
