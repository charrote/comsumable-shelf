"""Inventory API routes."""

from typing import Optional, List
from datetime import datetime
from sqlalchemy import select, update, cast, String
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from app.schemas import (
    InventoryResponse,
    TrackingReelResponse,
    InventoryUpdateRequest,
    InventoryUpdateResponse,
    DirectOutRequest,
    DirectOutResponse,
)
from app.utils.database import get_db
from app.models import InventoryReel, MaterialMaster, Shelf, ShelfSlot, Transaction, Customer
from app.services.inventory_service import direct_out
import openpyxl
from openpyxl.styles import Font, PatternFill
from urllib.parse import quote
import io

router = APIRouter(prefix="/inventory", tags=["Inventory"])


class ScanReelRequest(BaseModel):
    barcode: str


class ScanReelResponse(BaseModel):
    reel_id: int
    material_code: str
    material_name: Optional[str] = None
    quantity: float = 0
    shelf_code: Optional[str] = None
    status: str = ""


@router.post("/scan-reel", response_model=ScanReelResponse)
async def scan_reel_for_direct_out(
    data: ScanReelRequest,
    db: AsyncSession = Depends(get_db),
):
    """Scan a reel barcode to get info for direct outbound."""
    barcode = data.barcode.strip()
    if not barcode:
        raise HTTPException(status_code=400, detail="条码不能为空")

    reel_id = None
    try:
        reel_id = int(barcode)
    except ValueError:
        pass

    query = select(InventoryReel)
    if reel_id:
        query = query.where(InventoryReel.id == reel_id)
    else:
        query = query.where(InventoryReel.reel_barcode == barcode)

    result = await db.execute(query)
    reel = result.scalar_one_or_none()
    if not reel:
        raise HTTPException(status_code=404, detail="未找到该料盘")

    mat_result = await db.execute(
        select(MaterialMaster).where(MaterialMaster.id == reel.material_id)
    )
    material = mat_result.scalar_one_or_none()

    shelf_code = None
    if reel.shelf_slot_id:
        slot_result = await db.execute(
            select(Shelf.code)
            .select_from(ShelfSlot)
            .join(Shelf, ShelfSlot.shelf_id == Shelf.id)
            .where(ShelfSlot.id == reel.shelf_slot_id)
        )
        sc = slot_result.scalar_one_or_none()
        if sc:
            shelf_code = sc

    return ScanReelResponse(
        reel_id=reel.id,
        material_code=material.code if material else "",
        material_name=material.name if material else None,
        quantity=reel.quantity,
        shelf_code=shelf_code,
        status=reel.status,
    )


@router.get("")
async def get_inventory(
    customer_id: Optional[int] = Query(None, description="（已废弃，请使用 customer_ids）"),
    customer_ids: Optional[List[int]] = Query(None, description="按客户ID筛选（可多选）"),
    material_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None, description="（已废弃，请使用 statuses）"),
    statuses: Optional[List[str]] = Query(None, description="按状态筛选（可多选）: pending_shelving / on_shelf / in_use / tracking / exhausted"),
    keyword: Optional[str] = Query(None, description="搜索关键词（物料编号 / Reel 编码 / Reel ID）"),
    db: AsyncSession = Depends(get_db),
):
    """Query inventory pallets."""
    query = select(
        InventoryReel,
        MaterialMaster.code.label("material_code"),
        Shelf.code.label("shelf_code"),
        Customer.name.label("customer_name"),
        Customer.code.label("customer_code"),
    ).join(
        MaterialMaster, InventoryReel.material_id == MaterialMaster.id
    ).join(
        Customer, InventoryReel.customer_id == Customer.id
    ).outerjoin(
        ShelfSlot, InventoryReel.shelf_slot_id == ShelfSlot.id
    ).outerjoin(
        Shelf, ShelfSlot.shelf_id == Shelf.id
    )

    # Support both old single-value and new multi-value filters
    if customer_ids:
        query = query.where(InventoryReel.customer_id.in_(customer_ids))
    elif customer_id:
        query = query.where(InventoryReel.customer_id == customer_id)

    if material_id:
        query = query.where(InventoryReel.material_id == material_id)

    if statuses:
        query = query.where(InventoryReel.status.in_(statuses))
    elif status:
        query = query.where(InventoryReel.status == status)

    if keyword:
        keyword_like = f"%{keyword}%"
        query = query.where(
            MaterialMaster.code.ilike(keyword_like)
            | InventoryReel.reel_code.ilike(keyword_like)
            | InventoryReel.id.cast(String).ilike(keyword_like)
        )

    result = await db.execute(query)
    rows = result.all()

    pallets = []
    total_qty = 0
    exhausted = 0

    for row in rows:
        pallet = row[0]
        total_qty += pallet.quantity
        if pallet.status == "exhausted":
            exhausted += 1
        pallets.append({
            "reel_id": pallet.id,
            "material_code": row[1],
            "quantity": pallet.quantity,
            "original_quantity": pallet.original_quantity,
            "shelf_slot_id": pallet.shelf_slot_id,
            "shelf_code": row[2],
            "customer_name": row[3],
            "customer_code": row[4],
            "customer_id": pallet.customer_id,
            "first_in_time": pallet.first_in_time,
            "last_in_time": pallet.last_in_time,
            "status": pallet.status,
        })

    return InventoryResponse(
        pallets=pallets,
        summary={
            "total_pallets": len(pallets),
            "total_quantity": total_qty,
            "exhausted_pallets": exhausted,
        },
    )


@router.get("/tracking")
async def get_tracking_inventory(
    db: AsyncSession = Depends(get_db),
):
    """Get tracking inventory (returned items waiting for restock)."""
    result = await db.execute(
        select(InventoryReel, MaterialMaster.code.label("material_code"))
        .join(MaterialMaster, InventoryReel.material_id == MaterialMaster.id)
        .where(InventoryReel.status == "tracking")
    )
    rows = result.all()

    pallets = []
    for pallet, code in rows:
        pallets.append(TrackingReelResponse(
            reel_id=pallet.id,
            material_code=code,
            quantity=pallet.quantity,
            last_out_time=pallet.last_out_time,
            status=pallet.status,
        ))

    return {"pallets": pallets}


@router.put("/{reel_id}", response_model=InventoryUpdateResponse)
async def update_inventory_pallet(
    reel_id: int,
    data: InventoryUpdateRequest,
    db: AsyncSession = Depends(get_db),
):
    """Update inventory pallet fields (quantity, status, shelf_slot_id).

    Supports partial update — only provided fields are modified.
    Records a Transaction entry when quantity or status changes.
    """
    # 1. Verify pallet exists
    result = await db.execute(
        select(InventoryReel).where(InventoryReel.id == reel_id)
    )
    pallet = result.scalar_one_or_none()
    if not pallet:
        raise HTTPException(status_code=404, detail="库存托盘不存在")

    updated_fields = []
    old_quantity = pallet.quantity
    old_status = pallet.status

    # 2. Update quantity (if provided)
    if data.quantity is not None:
        pallet.quantity = data.quantity
        updated_fields.append("quantity")

    # 3. Update status (if provided)
    if data.status is not None:
        valid_statuses = {"pending_shelving", "on_shelf", "in_use", "tracking", "exhausted"}
        if data.status not in valid_statuses:
            raise HTTPException(
                status_code=400,
                detail=f"无效的状态值: {data.status}，有效值: {', '.join(sorted(valid_statuses))}",
            )
        pallet.status = data.status
        updated_fields.append("status")

    # 4. Update shelf_slot_id (if provided)
    if data.shelf_slot_id is not None:
        # Verify the slot exists
        slot_result = await db.execute(
            select(ShelfSlot).where(ShelfSlot.id == data.shelf_slot_id)
        )
        if not slot_result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail="储位不存在")

        # Check slot is not already occupied by a DIFFERENT pallet
        occupied = await db.execute(
            select(InventoryReel).where(
                InventoryReel.shelf_slot_id == data.shelf_slot_id,
                InventoryReel.id != reel_id,
                InventoryReel.status == "on_shelf",
            )
        )
        if occupied.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="该储位已被其它库存托盘占用")
        pallet.shelf_slot_id = data.shelf_slot_id
        updated_fields.append("shelf_slot_id")
    elif data.shelf_slot_id is None and "shelf_slot_id" in data.model_fields_set:
        # Explicitly set to null (unbind from slot)
        pallet.shelf_slot_id = None
        updated_fields.append("shelf_slot_id")

    # 5. Record transaction if quantity or status changed
    if data.quantity is not None or data.status is not None:
        now = datetime.utcnow()
        # Compute quantity diff for transaction
        qty_diff = data.quantity - old_quantity if data.quantity is not None else 0
        txn = Transaction(
            customer_id=pallet.customer_id,
            material_id=pallet.material_id,
            type="reverse" if qty_diff > 0 else "out" if qty_diff < 0 else "reverse",
            quantity=abs(qty_diff) if qty_diff != 0 else 0,
            balance_after=data.quantity if data.quantity is not None else old_quantity,
            reel_id=pallet.id,
            source_type="manual_adjust",
            note=data.note or f"手动调整: {', '.join(updated_fields)}",
            created_at=now,
        )
        db.add(txn)

    pallet.updated_at = datetime.utcnow()
    await db.commit()

    return InventoryUpdateResponse(
        status="ok",
        reel_id=pallet_id,
        updated_fields=updated_fields,
        message=f"库存托盘 #{reel_id} 已更新: {', '.join(updated_fields) if updated_fields else '无变更'}",
    )


@router.get("/export")
async def export_inventory(
    customer_ids: Optional[List[int]] = Query(None, description="按客户ID筛选（可多选）"),
    statuses: Optional[List[str]] = Query(None, description="按状态筛选（可多选）"),
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    db: AsyncSession = Depends(get_db),
):
    """Export inventory to Excel (.xlsx)."""
    query = select(
        InventoryReel,
        MaterialMaster.code.label("material_code"),
        MaterialMaster.name.label("material_name"),
        Shelf.code.label("shelf_code"),
        Customer.name.label("customer_name"),
        Customer.code.label("customer_code"),
    ).join(
        MaterialMaster, InventoryReel.material_id == MaterialMaster.id
    ).join(
        Customer, InventoryReel.customer_id == Customer.id
    ).outerjoin(
        ShelfSlot, InventoryReel.shelf_slot_id == ShelfSlot.id
    ).outerjoin(
        Shelf, ShelfSlot.shelf_id == Shelf.id
    )

    if customer_ids:
        query = query.where(InventoryReel.customer_id.in_(customer_ids))
    if statuses:
        query = query.where(InventoryReel.status.in_(statuses))
    if keyword:
        keyword_like = f"%{keyword}%"
        query = query.where(
            MaterialMaster.code.ilike(keyword_like)
            | InventoryReel.reel_code.ilike(keyword_like)
            | InventoryReel.id.cast(String).ilike(keyword_like)
        )

    result = await db.execute(query)
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存列表"

    # Headers
    headers = [
        "库存盘号", "物料编号", "物料名称", "数量", "原始数量",
        "客户名称", "客户编码", "储位编号", "储架编码",
        "状态", "首次入库时间", "最近入库时间"
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    status_labels = {
        "pending_shelving": "待上架", "on_shelf": "在架",
        "in_use": "使用中", "tracking": "跟踪中", "exhausted": "已耗尽",
    }

    for row in rows:
        pallet = row[0]
        ws.append([
            pallet.id,
            row[1],  # material_code
            row[2],  # material_name
            pallet.quantity,
            pallet.original_quantity,
            row[4],  # customer_name
            row[5],  # customer_code
            pallet.shelf_slot_id or "",
            row[3] or "",  # shelf_code
            status_labels.get(pallet.status, pallet.status),
            pallet.first_in_time.strftime("%Y-%m-%d %H:%M:%S") if pallet.first_in_time else "",
            pallet.last_in_time.strftime("%Y-%m-%d %H:%M:%S") if pallet.last_in_time else "",
        ])

    # Column widths
    widths = [14, 22, 30, 10, 12, 20, 16, 14, 14, 12, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"库存列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/reels/{reel_id}/direct-out", response_model=DirectOutResponse)
async def direct_outbound(
    reel_id: int,
    data: DirectOutRequest,
    db: AsyncSession = Depends(get_db),
):
    """Direct outbound — scan & release reel without BOM / IssueOrder.

    This is a simplified outbound flow for urgent picks or waste disposal.
    It reduces the reel's quantity, creates a Transaction record,
    and optionally releases the shelf slot when fully consumed.
    """
    result = await direct_out(
        db=db,
        reel_id=reel_id,
        quantity=data.quantity,
        operator=data.operator,
        note=data.note,
        release_slot=data.release_slot,
    )

    if result["status"] == "error":
        raise HTTPException(status_code=400, detail=result["message"])

    return DirectOutResponse(
        status=result["status"],
        reel_id=result["reel_id"],
        quantity_before=result["quantity_before"],
        quantity_after=max(0, result["quantity_after"]),
        reel_status=result["reel_status"],
        slot_released=result["slot_released"],
        message=result["message"],
    )
